import os
import openpyxl
import pandas as pd
import logging

logging.basicConfig(
    filename='D:\\Master\\Thesis\\CodeSmellsDetector\\ML_Approach\\ExtractedData\\LongMethod\\example.log',
    filemode='w', level=logging.DEBUG)
samples_path = 'D:\\Master\\Thesis\\CodeSmellsDetector\\ML_Approach\\ExtractedData\\LongMethod'
root = 'D:\\Master\\Thesis\\RDT\\C_Data\\m_r'
smell_name = 'Long Method'
positive_long_method_list = [['LOC', 'CC', 'PC', 'Is_Long_Method']]
negative_long_method_list = [['LOC', 'CC', 'PC', 'Is_Long_Method']]
positive_list = set()
negative_list = set()


def write_to_excel(path, list_list):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    for row_num, data in enumerate(list_list):
        ws.cell(row=row_num + 1, column=1).value = data[0]
        ws.cell(row=row_num + 1, column=2).value = data[1]
        ws.cell(row=row_num + 1, column=3).value = data[2]
        ws.cell(row=row_num + 1, column=4).value = data[3]

    wb.save(path)


def append_positive_negative_lists(smells_df, metrics_dict):
    long_method_metrics = set()

    for index, smell in smells_df.iterrows():
        try:
            key = str(smell['Project Name']) + "_" + str(smell['Package Name']) + "_" + str(smell['Type Name']) + "_" + str(
                smell['Method Name'])
            if key in long_method_metrics:
                #print("Key already found in long_method_metrics set " + key)
                continue

            metrics_list = metrics_dict.get(key)
            value = str(metrics_list[0]) + "_" + str(metrics_list[1]) + "_" + str(metrics_list[2])
            if smell['Implementation Smell'] == smell_name:
                long_method_metrics.add(key)
                if value in positive_list:
                    #print("Key already found in positive_list set " + value)
                    continue
                positive_long_method_list.append([metrics_list[0], metrics_list[1], metrics_list[2], True])
                positive_list.add(value)
                #print("Adding Key to positive_long_method_list " + value)
            else:
                if value in negative_list:
                    #print("Key already found in negative_list set " + value)
                    continue
                negative_long_method_list.append([metrics_list[0], metrics_list[1], metrics_list[2], False])
                negative_list.add(value)
                #print("Adding Key to negative_long_method_list " + value)
        except Exception as e:
            print("Error: Something went wrong for ", key, e)

    print("Positive long method Count: " + str(len(positive_long_method_list)))
    print("Negative long method Count: " + str(len(negative_long_method_list)))


def prepare_dict(metrics_df):
    metrics_dict = {}
    print("Preparing Dict...")
    for index, metrics in metrics_df.iterrows():
        key = str(metrics['Project Name']) + "_" + str(metrics['Package Name']) + "_" + str(
            metrics['Type Name']) + "_" + str(metrics['Method Name'])
        if key in metrics_dict:
            #print("Key Found " + key)
            logging.info('we have same method name here: ' + key)
        else:
            #print("Adding Key to Dict "+key)
            metrics_dict[key] = [metrics['LOC'], metrics['CC'], metrics['PC']]
    return metrics_dict


def create_long_method_dataset():
    dirlist = [item for item in os.listdir(root) if os.path.isdir(os.path.join(root, item))]
    for folder in dirlist:

        try:
            ###### Long Method ######
            out_dir = os.path.join(root, folder)

            imp_smells_path = out_dir + '\ImplementationSmells.csv'
            method_metrics_path = out_dir + '\MethodMetrics.csv'

            smells_data = pd.read_csv(imp_smells_path,encoding='cp1252')
            smells_df = pd.DataFrame(smells_data, columns=['Project Name', 'Package Name', 'Type Name', 'Method Name',
                                                           'Implementation Smell', 'Cause of the Smell'])

            metrics_data = pd.read_csv(method_metrics_path,encoding='cp1252')
            metrics_df = pd.DataFrame(metrics_data, columns=['Project Name', 'Package Name', 'Type Name', 'Method Name',
                                                             'LOC', 'CC', 'PC'])

            metrics_dict = prepare_dict(metrics_df)
            append_positive_negative_lists(smells_df, metrics_dict)

            print(folder, " is finished")
            print("------------------------------------------------------------------------")
            logging.info(folder + ' is finished')
            logging.info('------------------------------------------------------------------------')

        except Exception as e:
            print("Error: Something went wrong for ", folder, e)
        # create_long_parameter_list_dataset(smells_df,metrics_df)

    path1 = samples_path + '\\' + 'PositiveSamples.xlsx'
    path2 = samples_path + '\\' + 'NegativeSamples.xlsx'
    write_to_excel(path1, positive_long_method_list)
    write_to_excel(path2, negative_long_method_list)


create_long_method_dataset()
