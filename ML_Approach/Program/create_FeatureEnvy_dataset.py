import os
import openpyxl
import pandas as pd
import logging
import csv

from DataModels.MetricsModel import MetricsModel
from DataModels.ClassMetrics import ClassMetrics
from DataModels.Smell import Smell
from DataModels.SmellModel import SmellModel

logging.basicConfig(
    filename='D:\\Master\\Thesis\\CodeSmellsDetector\\ML_Approach\\ExtractedData\\FeatureEnvy\\log.log',
    filemode='w', level=logging.DEBUG)
samples_path = 'D:\\Master\\Thesis\\CodeSmellsDetector\\ML_Approach\\ExtractedData\\FeatureEnvy'
root = 'D:\\Master\\Thesis\\RDT\\C_Data\\m_r'
smell_name = 'Feature Envy'
positive_long_method_list = [
    ['NOF', 'NOPF', 'NOM', 'NOPM', 'LOC', 'WMC', 'NC', 'DIT', 'LCOM', 'FANIN', 'FANOUT', 'Is_Feature_Envy']]
negative_long_method_list = [
    ['NOF', 'NOPF', 'NOM', 'NOPM', 'LOC', 'WMC', 'NC', 'DIT', 'LCOM', 'FANIN', 'FANOUT', 'Is_Feature_Envy']]
positive_list = set()
negative_list = set()

metrics_dict = {}


# The tool detected a instance of this smell because stdCharacters is more interested in members of the type: EjbInfo
def prepare_metrics_dict(metrics_df):
    metrics_dict = {}
    print("Preparing metrics Dict...")
    for index, metrics in metrics_df.iterrows():
        key = str(metrics['Package Name']) + "_" + str(metrics['Type Name'])
        if key in metrics_dict:
            metrics_model = metrics_dict.get(key)
            metrics_model.metrics_list.append(
                ClassMetrics(metrics['NOF'], metrics['NOPF'], metrics['NOM'], metrics['NOPM'], metrics['LOC'],
                             metrics['WMC'], metrics['NC'], metrics['DIT'], metrics['LCOM'], metrics['FANIN'],
                             metrics['FANOUT']))
            logging.info('we have more than one smell for the same class: ' + key)
        else:
            metrics_model = get_metrics_model(metrics)
            metrics_dict[key] = metrics_model
    return metrics_dict


def get_metrics_model(metrics):
    metrics_list = []
    metrics_object = ClassMetrics(metrics['NOF'], metrics['NOPF'], metrics['NOM'], metrics['NOPM'], metrics['LOC'],
                                  metrics['WMC'], metrics['NC'], metrics['DIT'], metrics['LCOM'], metrics['FANIN'],
                                  metrics['FANOUT'])
    metrics_list.append(metrics_object)
    metrics_model = MetricsModel(metrics_list)

    return metrics_model


def prepare_smells_dict(smells_df):
    smells_dict = {}
    print("Preparing Smells Dict...")
    for index, smell in smells_df.iterrows():
        key = str(smell['Package Name']) + "_" + str(smell['Type Name'])
        if key in smells_dict:
            smell_model = smells_dict.get(key)
            smell_model.smells_list.append(Smell(smell['Design Smell'], smell['Cause of the Smell']))
            is_smelly = True if smell['Design Smell'] == smell_name else False
            smell_model.is_smelly = smell_model.is_smelly or is_smelly
            logging.info('we have more than one smell for the same class: ' + key)
        else:
            smell_model = get_smell_model(smell)
            smells_dict[key] = smell_model
    return smells_dict


def get_smell_model(smell):
    smells_list = []
    smell_object = Smell(smell['Design Smell'], smell['Cause of the Smell'])
    smells_list.append(smell_object)
    is_smelly = True if smell['Design Smell'] == smell_name else False
    smell_model = SmellModel(is_smelly, smells_list)

    return smell_model


def write_to_excel(path, list_list):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    for row_num, data in enumerate(list_list):
        ws.cell(row=row_num + 1, column=1).value = data[0]
        ws.cell(row=row_num + 1, column=2).value = data[1]
        ws.cell(row=row_num + 1, column=3).value = data[2]
        ws.cell(row=row_num + 1, column=4).value = data[3]
        ws.cell(row=row_num + 1, column=5).value = data[4]
        ws.cell(row=row_num + 1, column=6).value = data[5]
        ws.cell(row=row_num + 1, column=7).value = data[6]
        ws.cell(row=row_num + 1, column=8).value = data[7]
        ws.cell(row=row_num + 1, column=9).value = data[8]
        ws.cell(row=row_num + 1, column=10).value = data[9]
        ws.cell(row=row_num + 1, column=11).value = data[10]
        ws.cell(row=row_num + 1, column=12).value = data[11]

    wb.save(path)

def write_to_csv(path, list_list):
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        for row_num, data in enumerate(list_list):
            writer.writerow(data)
        f.close()

def is_FeatureEnvy_smell(cause_of_smell: str):
    if "instance" in cause_of_smell and "more interested in members of the type" in cause_of_smell:
        return True


def append_positive_negative_lists(smells_dict, metrics_dict):
    for smell_key in smells_dict:
        smell_value = smells_dict[smell_key]
        metrics_value = metrics_dict.get(smell_key)

        if smell_value.is_smelly:
            for smell in smell_value.smells_list:
                if smell.imp_smell == smell_name:

                    for metrics in metrics_value.metrics_list:
                        metrics_concatenated_value = "{0}_{1}_{2}_{3}_{4}_{5}_{6}_{7}_{8}_{9}_{10}".format(
                            str(metrics.NOF), str(metrics.NOPF), str(metrics.NOM), str(metrics.NOPM), str(metrics.LOC),
                            str(metrics.WMC), str(metrics.NC), str(metrics.DIT), str(metrics.LCOM), str(metrics.FANIN),
                            str(metrics.FANOUT))
                        if is_FeatureEnvy_smell(
                                smell.cause_of_smell) and metrics_concatenated_value not in positive_list:
                            positive_list.add(metrics_concatenated_value)
                            positive_long_method_list.append([metrics.NOF, metrics.NOPF, metrics.NOM, metrics.NOPM,
                                                              metrics.LOC, metrics.WMC, metrics.NC, metrics.DIT,
                                                              metrics.LCOM, metrics.FANIN, metrics.FANOUT, True])
        else:
            for metrics in metrics_value.metrics_list:
                metrics_concatenated_value = "{0}_{1}_{2}_{3}_{4}_{5}_{6}_{7}_{8}_{9}_{10}".format(
                    str(metrics.NOF), str(metrics.NOPF), str(metrics.NOM), str(metrics.NOPM), str(metrics.LOC),
                    str(metrics.WMC), str(metrics.NC), str(metrics.DIT), str(metrics.LCOM), str(metrics.FANIN),
                    str(metrics.FANOUT))
                if metrics_concatenated_value not in negative_list:
                    negative_list.add(metrics_concatenated_value)
                    negative_long_method_list.append([metrics.NOF, metrics.NOPF, metrics.NOM, metrics.NOPM,
                                                      metrics.LOC, metrics.WMC, metrics.NC, metrics.DIT,
                                                      metrics.LCOM, metrics.FANIN, metrics.FANOUT, False])


def create_feature_envy_dataset():
    dirlist = [item for item in os.listdir(root) if os.path.isdir(os.path.join(root, item))]
    for folder in dirlist:

        try:
            out_dir = os.path.join(root, folder)

            imp_smells_path = out_dir + '\DesignSmells.csv'
            method_metrics_path = out_dir + '\TypeMetrics.csv'

            smells_data = pd.read_csv(imp_smells_path, encoding='cp1252')
            smells_df = pd.DataFrame(smells_data, columns=['Project Name', 'Package Name', 'Type Name', 'Design Smell',
                                                           'Cause of the Smell'])

            metrics_data = pd.read_csv(method_metrics_path, encoding='cp1252')
            metrics_df = pd.DataFrame(metrics_data, columns=['Project Name', 'Package Name', 'Type Name',
                                                             'NOF', 'NOPF', 'NOM', 'NOPM', 'LOC', 'WMC', 'NC', 'DIT',
                                                             'LCOM', 'FANIN', 'FANOUT'])

            metrics_dict = prepare_metrics_dict(metrics_df)
            smells_dict = prepare_smells_dict(smells_df)
            append_positive_negative_lists(smells_dict, metrics_dict)

            print(folder, " is finished")
            print("------------------------------------------------------------------------")
            logging.info(folder + ' is finished')
            logging.info('------------------------------------------------------------------------')

        except Exception as e:
            print("Error: Something went wrong for ", folder, e)
            print(folder, " is finished with Exception")
            print("------------------------------------------------------------------------")

    path1 = samples_path + '\\' + 'PositiveSamples.xlsx'
    path2 = samples_path + '\\' + 'NegativeSamples.xlsx'
    write_to_excel(path1, positive_long_method_list)
    write_to_excel(path2, negative_long_method_list)

    list_lists = positive_long_method_list
    list_lists.extend(negative_long_method_list)
    path3 = samples_path + '\\' + 'FeatureEnvyDB.csv'
    write_to_csv(path3, list_lists)
