import os
import openpyxl
import pandas as pd
import logging
import csv

from DataModels.MetricsModel import MetricsModel
from DataModels.MethodMetrics import MethodMetrics
from DataModels.Smell import Smell
from DataModels.SmellModel import SmellModel

logging.basicConfig(
    filename='D:\\Master\\Thesis\\CodeSmellsDetector\\ML_Approach\\ExtractedData\\LongParameterList\\log.log',
    filemode='w', level=logging.DEBUG)
samples_path = 'D:\\Master\\Thesis\\CodeSmellsDetector\\ML_Approach\\ExtractedData\\LongParameterList'
root = 'D:\\Master\\Thesis\\RDT\\C_Data\\m_r'
smell_name = 'Long Parameter List'
positive_long_method_list = [['LOC', 'CC', 'PC', 'Is_Long_Parameter_List']]
negative_long_method_list = [['LOC', 'CC', 'PC', 'Is_Long_Parameter_List']]
positive_list = set()
negative_list = set()

metrics_dict = {}


# The method has 5 parameters.
def prepare_metrics_dict(metrics_df):
    metrics_dict = {}
    print("Preparing metrics Dict...")
    for index, metrics in metrics_df.iterrows():
        key = str(metrics['Package Name']) + "_" + str(metrics['Type Name']) + "_" + str(metrics['Method Name'])
        if key in metrics_dict:
            metrics_model = metrics_dict.get(key)
            metrics_model.metrics_list.append(MethodMetrics(metrics['LOC'], metrics['CC'], metrics['PC']))
            logging.info('we have same method name here: ' + key)
        else:
            metrics_model = get_metrics_model(metrics)
            metrics_dict[key] = metrics_model
    return metrics_dict


def get_metrics_model(metrics):
    metrics_list = []
    metrics_object = MethodMetrics(metrics['LOC'], metrics['CC'], metrics['PC'])
    metrics_list.append(metrics_object)
    metrics_model = MetricsModel(metrics_list)

    return metrics_model


def prepare_smells_dict(smells_df):
    smells_dict = {}
    print("Preparing Smells Dict...")
    for index, smell in smells_df.iterrows():
        key = str(smell['Package Name']) + "_" + str(smell['Type Name']) + "_" + str(smell['Method Name'])
        if key in smells_dict:
            smell_model = smells_dict.get(key)
            smell_model.smells_list.append(Smell(smell['Implementation Smell'], smell['Cause of the Smell']))
            is_smelly = True if smell['Implementation Smell'] == smell_name else False
            smell_model.is_smelly = smell_model.is_smelly or is_smelly
            logging.info('we have same method name here: ' + key)
        else:
            smell_model = get_smell_model(smell)
            smells_dict[key] = smell_model
    return smells_dict


def get_smell_model(smell):
    smells_list = []
    smell_object = Smell(smell['Implementation Smell'], smell['Cause of the Smell'])
    smells_list.append(smell_object)
    is_smelly = True if smell['Implementation Smell'] == smell_name else False
    smell_model = SmellModel(is_smelly, smells_list)

    return smell_model


def write_to_excel(path, list_list):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    for row_num, data in enumerate(list_list):
        ws.cell(row=row_num + 1, column=1).value = data[0]
        ws.cell(row=row_num + 1, column=2).value = data[1]
        ws.cell(row=row_num + 1, column=3).value = data[2]
        ws.cell(row=row_num + 1, column=4).value = data[3]

    wb.save(path)

def write_to_csv(path, list_list):
    with open(path,"w", newline="") as f:
        writer = csv.writer(f)
        for row_num, data in enumerate(list_list):
            writer.writerow(data)
        f.close()

def is_LongParameterList_smell(cause_of_smell: str, parameters: int):
    if "The method has" in cause_of_smell and "parameters" in cause_of_smell:
        # get corresponding method by parsing the cause of smell text
        parsed_cause_smell = [int(s) for s in cause_of_smell.split() if s.isdigit()][0]
        if parsed_cause_smell == parameters:
            return True


def append_positive_negative_lists(smells_dict, metrics_dict):
    for smell_key in smells_dict:
        smell_value = smells_dict[smell_key]
        metrics_value = metrics_dict.get(smell_key)

        if smell_value.is_smelly:
            for smell in smell_value.smells_list:
                if smell.imp_smell == smell_name:

                    for metrics in metrics_value.metrics_list:
                        metrics_concatenated_value = str(metrics.loc) + "_" + str(metrics.cc) + "_" + str(metrics.pc)
                        if is_LongParameterList_smell(smell.cause_of_smell, metrics.pc):
                            if metrics_concatenated_value not in positive_list:
                                positive_list.add(metrics_concatenated_value)
                                positive_long_method_list.append([metrics.loc, metrics.cc, metrics.pc, True])
        else:
            for metrics in metrics_value.metrics_list:
                metrics_concatenated_value = str(metrics.loc) + "_" + str(metrics.cc) + "_" + str(metrics.pc)
                if metrics_concatenated_value not in negative_list:
                    negative_list.add(metrics_concatenated_value)
                    negative_long_method_list.append([metrics.loc, metrics.cc, metrics.pc, False])


def create_long_parameters_list_dataset():
    dirlist = [item for item in os.listdir(root) if os.path.isdir(os.path.join(root, item))]
    for folder in dirlist:

        try:

            out_dir = os.path.join(root, folder)

            imp_smells_path = out_dir + '\ImplementationSmells.csv'
            method_metrics_path = out_dir + '\MethodMetrics.csv'

            smells_data = pd.read_csv(imp_smells_path, encoding='cp1252')
            smells_df = pd.DataFrame(smells_data, columns=['Project Name', 'Package Name', 'Type Name', 'Method Name',
                                                           'Implementation Smell', 'Cause of the Smell'])

            metrics_data = pd.read_csv(method_metrics_path, encoding='cp1252')
            metrics_df = pd.DataFrame(metrics_data, columns=['Project Name', 'Package Name', 'Type Name', 'Method Name',
                                                             'LOC', 'CC', 'PC'])

            metrics_dict = prepare_metrics_dict(metrics_df)
            smells_dict = prepare_smells_dict(smells_df)
            append_positive_negative_lists(smells_dict, metrics_dict)

            print(folder, " is finished")
            print("------------------------------------------------------------------------")
            logging.info(folder + ' is finished')
            logging.info('------------------------------------------------------------------------')

        except Exception as e:
            print("Error: Something went wrong for ", folder, e)
            print(folder, " is finished with Exception")
            print("------------------------------------------------------------------------")

    path1 = samples_path + '\\' + 'PositiveSamples.xlsx'
    path2 = samples_path + '\\' + 'NegativeSamples.xlsx'
    write_to_excel(path1, positive_long_method_list)
    write_to_excel(path2, negative_long_method_list)

    list_lists = positive_long_method_list
    list_lists.extend(negative_long_method_list)
    path3 = samples_path + '\\' + 'LongParameterListDB.csv'
    write_to_csv(path3, list_lists)

